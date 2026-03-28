"""
PDF Table Scraper → Excel
-------------------------
Extracts all tables from a selectable (copy-pasteable) PDF and saves them
to a formatted .xlsx file — one sheet per table, plus a Summary sheet.

Usage:
    python pdf_scraper.py <path_to_pdf> [output_excel.xlsx]

Examples:
    python pdf_scraper.py report.pdf
    python pdf_scraper.py report.pdf my_output.xlsx
"""

import sys
import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  EXTRACTION
# ─────────────────────────────────────────────

def extract_tables_from_pdf(pdf_path: str) -> list[dict]:
    """Return a list of dicts: {page, table_index, dataframe}"""
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        print(f"  PDF loaded: {total_pages} page(s)")

        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()

            for tbl_idx, raw_table in enumerate(tables, start=1):
                if not raw_table or len(raw_table) < 2:
                    continue  # skip empty or header-only tables

                # First row → column headers; clean None → empty string
                headers = [str(cell).strip() if cell else "" for cell in raw_table[0]]

                # De-duplicate blank headers (col_1, col_2, …)
                seen, clean_headers = {}, []
                for h in headers:
                    if h == "" or h in seen:
                        base = h or "col"
                        count = seen.get(base, 0) + 1
                        seen[base] = count
                        clean_headers.append(f"{base}_{count}")
                    else:
                        seen[h] = 1
                        clean_headers.append(h)

                rows = [
                    [str(cell).strip() if cell is not None else "" for cell in row]
                    for row in raw_table[1:]
                ]

                df = pd.DataFrame(rows, columns=clean_headers)
                df = df.dropna(how="all").reset_index(drop=True)

                results.append({
                    "page": page_num,
                    "table_index": tbl_idx,
                    "dataframe": df,
                    "sheet_name": f"P{page_num}_T{tbl_idx}",
                })
                print(f"  ✓ Page {page_num}, Table {tbl_idx}: "
                      f"{len(df)} rows × {len(df.columns)} cols")

    return results


# ─────────────────────────────────────────────
#  EXCEL FORMATTING HELPERS
# ─────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
ALT_ROW_FILL = PatternFill("solid", fgColor="EEF2F7")   # light blue-grey
BORDER_COLOR = "B0BEC5"

def _thin_border():
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)

def _format_sheet(ws, df: pd.DataFrame):
    """Apply professional formatting to a data sheet."""
    # ── Header row ──
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()

    # ── Data rows ──
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = _thin_border()

    # ── Column widths (auto-fit, capped) ──
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_values = df.iloc[:, col_idx - 1].astype(str).tolist() + [col_name]
        max_len = max(len(v) for v in col_values)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    # ── Freeze header row ──
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _add_summary_sheet(wb, tables: list[dict], pdf_path: str):
    """Insert a Summary sheet at position 0."""
    ws = wb.create_sheet("Summary", 0)

    title_font  = Font(bold=True, size=14, color="1F3864", name="Arial")
    header_font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    body_font   = Font(size=10, name="Arial")

    ws["A1"] = "PDF Table Extraction Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Source file: {os.path.basename(pdf_path)}"
    ws["A2"].font = Font(size=10, italic=True, name="Arial", color="555555")
    ws["A3"] = f"Total tables extracted: {len(tables)}"
    ws["A3"].font = body_font

    # ── Table of contents ──
    headers = ["Sheet", "Page", "Table #", "Rows", "Columns"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    for row_idx, t in enumerate(tables, start=6):
        df = t["dataframe"]
        values = [t["sheet_name"], t["page"], t["table_index"], len(df), len(df.columns)]
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = body_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _thin_border()

    col_widths = [15, 8, 10, 8, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A6"


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def scrape_pdf_to_excel(pdf_path: str, output_path: str | None = None) -> str:
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    if output_path is None:
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        output_path = f"{base}_extracted.xlsx"

    print(f"\n📄 Extracting tables from: {pdf_path}")
    tables = extract_tables_from_pdf(pdf_path)

    if not tables:
        print("\n⚠️  No tables found in this PDF.")
        print("    Make sure the PDF is selectable (not a scanned image).")
        return ""

    print(f"\n📊 Writing {len(tables)} table(s) to Excel …")

    # Write data sheets via pandas first
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for t in tables:
            t["dataframe"].to_excel(writer, sheet_name=t["sheet_name"], index=False)

    # Re-open with openpyxl for formatting
    wb = load_workbook(output_path)
    for t in tables:
        ws = wb[t["sheet_name"]]
        _format_sheet(ws, t["dataframe"])

    _add_summary_sheet(wb, tables, pdf_path)
    wb.save(output_path)

    print(f"\n✅ Done! Saved to: {output_path}")
    print(f"   Sheets: Summary + {', '.join(t['sheet_name'] for t in tables)}")
    return output_path


# ─────────────────────────────────────────────
#  CLI ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    pdf_file    = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    scrape_pdf_to_excel(pdf_file, output_file)
